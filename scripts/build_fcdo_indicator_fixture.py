#!/usr/bin/env python3
"""One-off builder for FCDO indicator_data test workbook."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment

REPO = Path(__file__).resolve().parents[1]
OUT = (
    REPO
    / "tests"
    / "fixtures"
    / "indicator_extractor"
    / "fcdo_bridgelight_indicator_data.xlsx"
)


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Indicators"

    headers = [
        "row_id",
        "indicator_ref",
        "indicator_name",
        "target",
        "actual",
        "unit",
        "disagg_dimension",
        "disagg_total",
        "disagg_male",
        "disagg_female",
        "disagg_other",
    ]
    ws.append(headers)

    rows = [
        [
            "op1_1_girls_reenrolled",
            "OP1.1",
            "Girls re-enrolled to formal education",
            1200,
            985,
            "persons",
            "",
            "",
            "",
            "",
            "",
        ],
        [
            "op1_2_girls_attending",
            "OP1.2",
            "Girls attending at least 80% of sessions",
            900,
            712,
            "persons",
            "",
            "",
            "",
            "",
            "",
        ],
        [
            "disagg_non_sum",
            "DISAGG",
            "Beneficiaries reached by gender",
            100,
            98,
            "persons",
            "gender",
            100,
            40,
            35,
            30,
        ],
        [
            "op1_1_target_only",
            "OP1.T",
            "Girls trained Q4 (target only row)",
            500,
            None,
            "persons",
            "",
            "",
            "",
            "",
            "",
        ],
        ["", "", "", "", "", "", "", "", "", "", ""],
        [
            "hidden_continuation_row",
            "OP-HIDDEN",
            "District learning meetings held (continuation block)",
            12,
            11,
            "count",
            "",
            "",
            "",
            "",
            "",
        ],
        [
            "cell_state_demo",
            "DEMO",
            "Cell state demonstration row",
            0,
            None,
            "N/A",
            "",
            "",
            "",
            "",
            "",
        ],
        [
            "op2_1_latrine_stances",
            "OP2.1",
            "Latrine stances rehabilitated",
            24,
            22,
            "stances",
            "",
            "",
            "",
            "",
            "",
        ],
        [
            "ocm1_attendance_80pct",
            "OCM1",
            "Girls meeting 80% attendance threshold",
            "70%",
            "68%",
            "percent",
            "",
            "",
            "",
            "",
            "",
        ],
    ]
    for row in rows:
        ws.append(row)

    ws.merge_cells("A6:C6")
    ws["A6"] = ""
    ws["A6"].alignment = Alignment(horizontal="center")

    fin = wb.create_sheet("Financials")
    fin.append(["line", "label", "budget", "actual", "currency"])
    fin.append(
        [
            "budget_total",
            "Total programme budget",
            1240000,
            1184000,
            "GBP",
        ]
    )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
