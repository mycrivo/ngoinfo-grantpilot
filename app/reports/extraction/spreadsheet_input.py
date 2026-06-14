"""Direct spreadsheet intake for indicator_data documents.

Preserves cell state (stated / blank / not_applicable) and row identity.
Docling markdown is not used for indicator_data — see ME_MODULE D-036.
P5: `.docx` monitoring tables via Docling table export (layout-aware read).
"""

from __future__ import annotations

import csv
import hashlib
import json
import math
import re
from pathlib import Path
from typing import Any, Literal

CellState = Literal["stated", "blank", "not_applicable"]

_NA_VALUES = frozenset({"n/a", "na", "n.a.", "-"})


def _classify_cell_raw(raw: str | None) -> tuple[str | None, CellState]:
    if raw is None:
        return None, "blank"
    text = str(raw).strip()
    if text == "":
        return None, "blank"
    if text.lower() in _NA_VALUES:
        return text, "not_applicable"
    return text, "stated"


def _format_cell_ref(col: int, row: int) -> str:
    from openpyxl.utils import get_column_letter

    return f"{get_column_letter(col)}{row}"


def _normalize_tabular_cell_value(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return str(value)
    text = str(value).strip()
    return text or None


def _normalize_xlsx_value(value: Any) -> str | None:
    return _normalize_tabular_cell_value(value)


def _rows_from_dataframe(df: Any) -> list[dict[str, Any]]:
    """Convert a Docling/pandas table to the shared indicator grid shape."""
    rows: list[dict[str, Any]] = []

    header_cells: list[dict[str, Any]] = []
    for col_idx, col_name in enumerate(df.columns, start=1):
        raw, cell_state = _classify_cell_raw(_normalize_tabular_cell_value(col_name))
        header_cells.append(
            {
                "ref": _format_cell_ref(col_idx, 1),
                "raw": raw,
                "cell_state": cell_state,
            }
        )
    rows.append({"row_index": 1, "cells": header_cells})

    for offset, (_, series) in enumerate(df.iterrows(), start=2):
        cells: list[dict[str, Any]] = []
        for col_idx, value in enumerate(series.tolist(), start=1):
            raw_text = _normalize_tabular_cell_value(value)
            raw, cell_state = _classify_cell_raw(raw_text)
            cells.append(
                {
                    "ref": _format_cell_ref(col_idx, offset),
                    "raw": raw,
                    "cell_state": cell_state,
                }
            )
        rows.append({"row_index": offset, "cells": cells})
    return rows


def parse_docx_tables(path: Path) -> dict[str, Any]:
    """Extract Word table grids via Docling — values read from cells, never inferred."""
    from docling.document_converter import DocumentConverter

    result = DocumentConverter().convert(str(path))
    document = result.document
    if not document.tables:
        raise ValueError(f"No tables found in document: {path.name}")

    sheets: list[dict[str, Any]] = []
    for table_idx, table in enumerate(document.tables):
        dataframe = table.export_to_dataframe(doc=document)
        sheet_name = f"Table{table_idx + 1}"
        sheets.append(
            {
                "name": sheet_name,
                "rows": _rows_from_dataframe(dataframe),
            }
        )

    return {
        "source_path": str(path),
        "format": "docx",
        "sheets": sheets,
    }


def parse_xlsx_workbook(path: Path) -> dict[str, Any]:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True, read_only=False)
    sheets: list[dict[str, Any]] = []
    for ws in wb.worksheets:
        rows: list[dict[str, Any]] = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=1), start=1):
            cells: list[dict[str, Any]] = []
            for col_idx, cell in enumerate(row, start=1):
                raw_text = _normalize_xlsx_value(cell.value)
                raw, cell_state = _classify_cell_raw(raw_text)
                cells.append(
                    {
                        "ref": _format_cell_ref(col_idx, row_idx),
                        "raw": raw,
                        "cell_state": cell_state,
                    }
                )
            rows.append({"row_index": row_idx, "cells": cells})
        sheets.append({"name": ws.title, "rows": rows})
    wb.close()
    return {
        "source_path": str(path),
        "format": "xlsx",
        "sheets": sheets,
    }


def parse_csv_file(path: Path) -> dict[str, Any]:
    rows: list[dict[str, Any]] = []
    with path.open(newline="", encoding="utf-8") as handle:
        reader = csv.reader(handle)
        for row_idx, row_values in enumerate(reader, start=1):
            cells: list[dict[str, Any]] = []
            for col_idx, value in enumerate(row_values, start=1):
                raw, cell_state = _classify_cell_raw(value if value != "" else None)
                cells.append(
                    {
                        "ref": _format_cell_ref(col_idx, row_idx),
                        "raw": raw,
                        "cell_state": cell_state,
                    }
                )
            rows.append({"row_index": row_idx, "cells": cells})
    return {
        "source_path": str(path),
        "format": "csv",
        "sheets": [{"name": path.stem, "rows": rows}],
    }


def parse_spreadsheet_from_path(path: Path) -> dict[str, Any]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return parse_xlsx_workbook(path)
    if suffix == ".csv":
        return parse_csv_file(path)
    if suffix == ".docx":
        return parse_docx_tables(path)
    raise ValueError(f"Unsupported spreadsheet format: {suffix}")


def spreadsheet_to_json_text(data: dict[str, Any], *, max_chars: int = 120_000) -> tuple[str, bool]:
    text = json.dumps(data, ensure_ascii=False)
    if len(text) <= max_chars:
        return text, False
    return text[:max_chars], True


def compute_spreadsheet_hash(data: dict[str, Any]) -> str:
    normalized = json.dumps(data, sort_keys=True, ensure_ascii=False)
    return hashlib.sha256(normalized.encode("utf-8")).hexdigest()


def get_cell_at_ref(data: dict[str, Any], sheet_name: str, cell_ref: str) -> dict[str, Any] | None:
    ref = cell_ref
    if "!" in ref:
        sheet_part, ref = ref.split("!", 1)
        sheet_name = sheet_part
    ref = ref.upper()
    col_letters = re.match(r"([A-Z]+)", ref)
    row_digits = re.search(r"(\d+)$", ref)
    if not col_letters or not row_digits:
        return None
    from openpyxl.utils import column_index_from_string

    col = column_index_from_string(col_letters.group(1))
    row = int(row_digits.group(1))
    sheet_name_lower = sheet_name.lower()
    for sheet in data.get("sheets", []):
        if (sheet.get("name") or "").lower() != sheet_name_lower:
            continue
        for row_data in sheet.get("rows", []):
            if row_data.get("row_index") != row:
                continue
            for cell in row_data.get("cells", []):
                cell_col = column_index_from_string(
                    re.match(r"([A-Z]+)", cell["ref"]).group(1)  # type: ignore[union-attr]
                )
                if cell_col == col:
                    return cell
    return None


# Package A: header phrases that identify a funder's source-declared section column
# (e.g. NLCF monitoring "Section for NLCF update"). Markdown-stripped, lower-cased,
# substring match. Kept deliberately narrow so ordinary indicator headers
# ("indicator", "target", "actual", "evidence", demographics) never match.
_SECTION_COLUMN_HEADER_TOKENS: tuple[str, ...] = ("section",)


def _normalize_header(raw: str | None) -> str:
    if raw is None:
        return ""
    return re.sub(r"[*_`]+", "", str(raw)).strip().lower()


def _column_letters(ref: str | None) -> str | None:
    if not ref:
        return None
    match = re.match(r"([A-Z]+)", str(ref).upper())
    return match.group(1) if match else None


def locate_section_assignment_column(
    data: dict[str, Any],
) -> dict[str, dict[int, dict[str, str]]]:
    """Locate the source-declared section column per sheet (Package A carrier).

    Deterministic, header-driven: finds the column whose header cell text matches a
    section-column token, then reads that column's verbatim value for every data row.
    Returns ``{sheet_name: {row_index: {"raw": label, "cell_ref": "Sheet!A<row>"}}}``.

    A sheet with no matching header yields no entry — observable as an absent sheet so
    the caller can surface the gap rather than silently routing to None. Values are
    copied verbatim from the grid; section membership is never inferred.
    """
    out: dict[str, dict[int, dict[str, str]]] = {}
    for sheet in data.get("sheets", []):
        sheet_name = sheet.get("name") or ""
        rows = sheet.get("rows") or []
        if not rows:
            continue
        header_cells = rows[0].get("cells") or []
        section_col: str | None = None
        for cell in header_cells:
            normalized = _normalize_header(cell.get("raw"))
            if normalized and any(tok in normalized for tok in _SECTION_COLUMN_HEADER_TOKENS):
                section_col = _column_letters(cell.get("ref"))
                break
        if section_col is None:
            continue
        per_row: dict[int, dict[str, str]] = {}
        for row_data in rows[1:]:
            row_index = row_data.get("row_index")
            if row_index is None:
                continue
            for cell in row_data.get("cells") or []:
                if _column_letters(cell.get("ref")) != section_col:
                    continue
                raw = cell.get("raw")
                if raw is None or str(raw).strip() == "":
                    break
                per_row[int(row_index)] = {
                    "raw": str(raw).strip(),
                    "cell_ref": f"{sheet_name}!{section_col}{row_index}",
                }
                break
        if per_row:
            out[sheet_name] = per_row
    return out


def list_data_row_ids(data: dict[str, Any], *, sheet_name: str = "Indicators") -> list[str]:
    """Return row_id values from column A (first cell) for rows after the header."""
    for sheet in data.get("sheets", []):
        if sheet.get("name") != sheet_name:
            continue
        row_ids: list[str] = []
        for row_data in sheet.get("rows", [])[1:]:
            cells = row_data.get("cells") or []
            if not cells:
                continue
            raw = cells[0].get("raw")
            if raw:
                row_ids.append(str(raw))
        return row_ids
    return []
